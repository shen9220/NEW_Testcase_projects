import os
import json
import zipfile
from app.config import settings
from app.utils.file_utils import ensure_dir

# Minimal META-INF/manifest.xml required by XMind desktop app to open .xmind files.
# The xmind library omits this when creating new files from scratch.
_MANIFEST_XML = """<?xml version="1.0" encoding="UTF-8" standalone="no"?>
<manifest xmlns="urn:xmind:xmap:xmlns:manifest:2.0" password-hint="">
<file-entry full-path="content.xml" media-type="text/xml"/>
<file-entry full-path="META-INF/" media-type=""/>
<file-entry full-path="META-INF/manifest.xml" media-type="text/xml"/>
</manifest>"""


def _inject_manifest(filepath: str) -> None:
    """Re-pack the XMind ZIP to include the missing META-INF/manifest.xml."""
    tmp_path = filepath + ".tmp"
    with zipfile.ZipFile(filepath, 'r') as zin:
        with zipfile.ZipFile(tmp_path, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                zout.writestr(item, zin.read(item.filename))
            zout.writestr("META-INF/manifest.xml", _MANIFEST_XML)
    os.replace(tmp_path, filepath)


def generate_excel(testcases: list, task_id: str, module: str = "") -> str:
    """Generate .xlsx file from test cases. Returns file path."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"

    # Header style
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["编号", "模块", "用例标题", "前置条件", "测试步骤", "预期结果", "优先级", "类型", "标签", "备注"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Data rows
    for row_idx, tc in enumerate(testcases, 2):
        steps_action = "\n".join([f"{i + 1}. {s.get('action', '')}" for i, s in enumerate(tc.get("steps", []))])
        steps_expected = "\n".join([f"{i + 1}. {s.get('expected', '')}" for i, s in enumerate(tc.get("steps", []))])
        tags_str = ", ".join(tc.get("tags", []))

        row_data = [
            tc.get("case_id", ""),
            tc.get("module", ""),
            tc.get("title", ""),
            tc.get("precondition", ""),
            steps_action,
            steps_expected,
            tc.get("priority", ""),
            tc.get("type", ""),
            tags_str,
            tc.get("notes", ""),
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Column widths
    widths = [8, 12, 30, 25, 50, 50, 8, 10, 15, 25]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    # Save
    suffix = f"_{module}" if module else ""
    export_dir = os.path.join(settings.data_dir, task_id, "export")
    ensure_dir(export_dir)
    filepath = os.path.join(export_dir, f"testcases{suffix}.xlsx")
    wb.save(filepath)
    return filepath


def generate_xmind(testcases: list, task_id: str, module: str = "") -> str:
    """Generate .xmind file organized by module → title. Returns file path."""
    try:
        import xmind
    except ImportError:
        raise ImportError("请安装 xmind 库: pip install xmind")

    suffix = f"_{module}" if module else ""
    export_dir = os.path.join(settings.data_dir, task_id, "export")
    ensure_dir(export_dir)
    filepath = os.path.join(export_dir, f"testcases{suffix}.xmind")

    # xmind.load() creates a new workbook when the file doesn't exist
    workbook = xmind.load(filepath)
    sheet = workbook.getPrimarySheet()
    sheet.setTitle("测试用例")
    root = sheet.getRootTopic()
    root.setTitle("测试用例")
    root.setStructureClass("org.xmind.ui.logic.right")

    # Group by module
    modules = {}
    for tc in testcases:
        mod = tc.get("module", "未分类")
        if mod not in modules:
            modules[mod] = []
        modules[mod].append(tc)

    for mod_name, cases in modules.items():
        mod_topic = root.addSubTopic()
        mod_topic.setTitle(f"{mod_name} ({len(cases)}条)")

        for tc in cases:
            case_topic = mod_topic.addSubTopic()
            case_topic.setTitle(f"[{tc.get('priority', '')}] {tc.get('title', '')}")

            prec = tc.get('precondition', '')
            if prec and prec != '无':
                pre_topic = case_topic.addSubTopic()
                pre_topic.setTitle(f"前置条件: {prec}")

            for i, step in enumerate(tc.get("steps", []), 1):
                step_topic = case_topic.addSubTopic()
                step_topic.setTitle(f"步骤{i}")
                action_topic = step_topic.addSubTopic()
                action_topic.setTitle(f"操作: {step.get('action', '')}")
                exp_topic = step_topic.addSubTopic()
                exp_topic.setTitle(f"预期结果: {step.get('expected', '')}")

            if tc.get("notes"):
                notes_topic = case_topic.addSubTopic()
                notes_topic.setTitle(f"备注: {tc['notes']}")

    xmind.save(workbook, filepath)
    _inject_manifest(filepath)
    return filepath
